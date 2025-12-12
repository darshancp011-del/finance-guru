import os
import matplotlib.pyplot as plt
import tempfile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import mimetypes
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
import mysql.connector
from db import get_db_connection
import hashlib
from datetime import date, datetime
import pandas as pd
from io import BytesIO
import tempfile
from fpdf import FPDF

# Ensure soft-delete column exists for transactions
def ensure_transaction_soft_delete():
    conn = get_db_connection()
    if conn is None:
        print("Warning: Could not connect to database for ensure_transaction_soft_delete")
        return
    cursor = conn.cursor()
    try:
        cursor.execute("ALTER TABLE transactions ADD COLUMN is_deleted BOOLEAN DEFAULT FALSE")
        conn.commit()
    except mysql.connector.Error:
        conn.rollback()  # Column likely already exists
    finally:
        cursor.close()
        conn.close()

# Ensure user profile columns exist (phone, job_title, bio, initial_balance)
def ensure_user_profile_columns():
    conn = get_db_connection()
    if conn is None:
        print("Warning: Could not connect to database for ensure_user_profile_columns")
        return
    cursor = conn.cursor()
    columns_to_add = [
        ("phone", "VARCHAR(20) DEFAULT NULL"),
        ("job_title", "VARCHAR(100) DEFAULT NULL"),
        ("bio", "TEXT DEFAULT NULL"),
        ("initial_balance", "DECIMAL(12, 2) DEFAULT 0.00")
    ]
    for col_name, col_def in columns_to_add:
        try:
            cursor.execute(f"ALTER TABLE users ADD COLUMN {col_name} {col_def}")
            conn.commit()
        except mysql.connector.Error:
            conn.rollback()  # Column likely already exists
    cursor.close()
    conn.close()

# Fix for Windows CSS MIME type issue
mimetypes.add_type('text/css', '.css')
mimetypes.add_type('application/javascript', '.js')

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# --- Components ---

def get_user_by_email(email):
    conn = get_db_connection()
    if conn is None:
        print("Error: Could not connect to database in get_user_by_email")
        return None
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE email = %s", (email,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()
    return user

def create_user(username, email, password, initial_balance=0):
    hash_pwd = hashlib.sha256(password.encode()).hexdigest()
    conn = get_db_connection()
    if conn is None:
        print("Error: Could not connect to database in create_user")
        return False
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO users (username, email, password_hash, initial_balance) VALUES (%s, %s, %s, %s)",
                       (username, email, hash_pwd, initial_balance))
        conn.commit()
        
        # Get the new user ID and create welcome notification
        user_id = cursor.lastrowid
        welcome_msg = f"🎉 Welcome to Finance Guru, {username}! Start tracking your finances today."
        cursor.execute("INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, 'success')",
                       (user_id, welcome_msg))
        
        # Add initial balance notification
        if initial_balance > 0:
            balance_msg = f"💰 Your starting balance of ₹{initial_balance:,.2f} has been set."
            cursor.execute("INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, 'info')",
                           (user_id, balance_msg))
        
        conn.commit()
        
        return True
    except mysql.connector.Error as err:
        print(f"Database error in create_user: {err}")
        return False
    finally:
        cursor.close()
        conn.close()

# --- Routes ---

app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'static', 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Initialize database and create tables at startup
from db import init_db
try:
    init_db()
    print("Database initialized successfully!")
except Exception as e:
    print(f"Warning: Could not initialize database: {e}")

# Ensure the soft-delete column exists at startup
ensure_transaction_soft_delete()

# Ensure user profile columns exist at startup
ensure_user_profile_columns()

@app.route('/health')
def health_check():
    """Health check endpoint to verify database connection"""
    import os
    try:
        conn = get_db_connection()
        if conn is None:
            return jsonify({
                'status': 'error',
                'message': 'Could not connect to database',
                'env_vars': {
                    'DB_HOST': os.environ.get('DB_HOST', 'NOT SET'),
                    'DB_USER': os.environ.get('DB_USER', 'NOT SET'),
                    'DB_NAME': os.environ.get('DB_NAME', 'NOT SET'),
                    'DB_PASSWORD': '***' if os.environ.get('DB_PASSWORD') else 'NOT SET'
                }
            }), 500
        
        cursor = conn.cursor()
        cursor.execute("SHOW TABLES")
        tables = [t[0] for t in cursor.fetchall()]
        cursor.close()
        conn.close()
        
        return jsonify({
            'status': 'ok',
            'database': 'connected',
            'tables': tables,
            'env_vars': {
                'DB_HOST': os.environ.get('DB_HOST', 'NOT SET'),
                'DB_USER': os.environ.get('DB_USER', 'NOT SET'),
                'DB_NAME': os.environ.get('DB_NAME', 'NOT SET'),
                'DB_PASSWORD': '***' if os.environ.get('DB_PASSWORD') else 'NOT SET'
            }
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/')
def home():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

# Store password reset tokens (in production, use database or Redis)
password_reset_tokens = {}

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        user = get_user_by_email(email)
        
        if user and user['password_hash'] == hashlib.sha256(password.encode()).hexdigest():
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['profile_pic'] = user.get('profile_pic', 'default.png')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password', 'error')
            
    return render_template('login.html')

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form['email']
        user = get_user_by_email(email)
        
        if user:
            # Generate a reset token
            import secrets
            token = secrets.token_urlsafe(32)
            password_reset_tokens[token] = {
                'email': email,
                'expires': datetime.now().timestamp() + 3600  # 1 hour expiry
            }
            
            # In production, send email here
            # For now, we'll show the reset link (remove in production!)
            reset_link = url_for('reset_password', token=token, _external=True)
            
            # Try to send email (if configured)
            try:
                send_reset_email(email, reset_link, user['username'])
                flash('Password reset link has been sent to your email!', 'success')
            except Exception as e:
                # If email fails, show the link directly (for development)
                flash(f'Reset link (dev mode): {reset_link}', 'info')
        else:
            # Don't reveal if email exists or not for security
            flash('If an account with that email exists, a reset link has been sent.', 'info')
        
        return redirect(url_for('forgot_password'))
    
    return render_template('forgot_password.html')

@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    # Check if token exists and is valid
    if token not in password_reset_tokens:
        flash('Invalid or expired reset link.', 'error')
        return redirect(url_for('forgot_password'))
    
    token_data = password_reset_tokens[token]
    
    # Check if token has expired
    if datetime.now().timestamp() > token_data['expires']:
        del password_reset_tokens[token]
        flash('Reset link has expired. Please request a new one.', 'error')
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        if new_password != confirm_password:
            flash('Passwords do not match.', 'error')
            return redirect(url_for('reset_password', token=token))
        
        if len(new_password) < 6:
            flash('Password must be at least 6 characters.', 'error')
            return redirect(url_for('reset_password', token=token))
        
        # Update password
        email = token_data['email']
        hash_pwd = hashlib.sha256(new_password.encode()).hexdigest()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET password_hash = %s WHERE email = %s", (hash_pwd, email))
        conn.commit()
        cursor.close()
        conn.close()
        
        # Remove used token
        del password_reset_tokens[token]
        
        flash('Password reset successful! Please login with your new password.', 'success')
        return redirect(url_for('login'))
    
    return render_template('reset_password.html', token=token)

def send_reset_email(email, reset_link, username):
    """Send password reset email using SMTP"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    # Email configuration (you can move these to environment variables)
    SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
    SMTP_PORT = int(os.environ.get('SMTP_PORT', 587))
    SMTP_USERNAME = os.environ.get('SMTP_USERNAME', '')
    SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
    
    if not SMTP_USERNAME or not SMTP_PASSWORD:
        raise Exception("Email not configured")
    
    msg = MIMEMultipart('alternative')
    msg['Subject'] = 'Finance Guru - Password Reset'
    msg['From'] = SMTP_USERNAME
    msg['To'] = email
    
    html = f"""
    <html>
    <body style="font-family: Arial, sans-serif; background-color: #0F172A; color: #F8FAFC; padding: 20px;">
        <div style="max-width: 500px; margin: 0 auto; background: rgba(255,255,255,0.05); border-radius: 16px; padding: 30px;">
            <h2 style="color: #7269E3;">Finance Guru</h2>
            <p>Hi {username},</p>
            <p>You requested to reset your password. Click the button below to create a new password:</p>
            <p style="text-align: center; margin: 30px 0;">
                <a href="{reset_link}" style="background: linear-gradient(135deg, #7269E3, #5b50d6); color: white; padding: 12px 30px; border-radius: 8px; text-decoration: none; font-weight: bold;">Reset Password</a>
            </p>
            <p style="color: #94A3B8; font-size: 14px;">This link will expire in 1 hour.</p>
            <p style="color: #94A3B8; font-size: 14px;">If you didn't request this, please ignore this email.</p>
        </div>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(html, 'html'))
    
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.sendmail(SMTP_USERNAME, email, msg.as_string())

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        initial_balance = request.form.get('initial_balance', 0)
        
        # Convert initial balance to float
        try:
            initial_balance = float(initial_balance) if initial_balance else 0
        except ValueError:
            initial_balance = 0
        
        if password != confirm_password:
            flash('Passwords do not match', 'error')
        else:
            if create_user(username, email, password, initial_balance):
                flash('Registration successful! Please login.', 'success')
                return redirect(url_for('login'))
            else:
                flash('Email already registered', 'error')
                
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/profile', methods=['GET', 'POST'])
def profile():
    if 'user_id' not in session:
        return redirect(url_for('login'))
        
    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        phone = request.form.get('phone', '')
        job_title = request.form.get('job_title', '')
        bio = request.form.get('bio', '')
        file = request.files.get('profile_pic')
        
        try:
            # Handle File Upload
            if file and file.filename != '':
                from werkzeug.utils import secure_filename
                filename = secure_filename(file.filename)
                unique_filename = f"{user_id}_{int(datetime.now().timestamp())}_{filename}"
                # Save new file
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
                # Fetch old profile picture to delete later
                cursor.execute("SELECT profile_pic FROM users WHERE id = %s", (user_id,))
                old_user = cursor.fetchone()
                old_pic = old_user['profile_pic'] if old_user else None
                # Update DB with new picture
                cursor.execute("UPDATE users SET profile_pic = %s WHERE id = %s", (unique_filename, user_id))
                session['profile_pic'] = unique_filename
                # Delete old picture file if it's a custom upload (not default) and different from new
                if old_pic and old_pic != 'default.png' and old_pic != unique_filename:
                    old_path = os.path.join(app.config['UPLOAD_FOLDER'], old_pic)
                    try:
                        if os.path.isfile(old_path):
                            os.remove(old_path)
                    except Exception as e:
                        print(f"Failed to delete old profile picture {old_path}: {e}")
            
            # Update Details
            cursor.execute("""
                UPDATE users 
                SET username = %s, email = %s, phone = %s, job_title = %s, bio = %s 
                WHERE id = %s
            """, (username, email, phone, job_title, bio, user_id))
            conn.commit()
            
            session['username'] = username
            flash('Profile updated successfully!', 'success')
            
        except mysql.connector.Error as err:
            flash(f'Error updating profile: {err}', 'error')
        
        return redirect(url_for('profile'))
    
    cursor.execute("SELECT * FROM users WHERE id = %s", (user_id,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()
    
    return render_template('profile.html', user=user)

@app.route('/change_password', methods=['POST'])
def change_password():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    current_password = request.form.get('current_password', '')
    new_password = request.form.get('new_password', '')
    confirm_password = request.form.get('confirm_password', '')
    
    # Validate inputs
    if not current_password or not new_password or not confirm_password:
        flash('All password fields are required', 'error')
        return redirect(url_for('profile'))
    
    if new_password != confirm_password:
        flash('New passwords do not match', 'error')
        return redirect(url_for('profile'))
    
    if len(new_password) < 6:
        flash('Password must be at least 6 characters', 'error')
        return redirect(url_for('profile'))
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Verify current password
    cursor.execute("SELECT password_hash FROM users WHERE id = %s", (user_id,))
    user = cursor.fetchone()
    
    if not user:
        cursor.close()
        conn.close()
        flash('User not found', 'error')
        return redirect(url_for('profile'))
    
    current_hash = hashlib.sha256(current_password.encode()).hexdigest()
    if user['password_hash'] != current_hash:
        cursor.close()
        conn.close()
        flash('Current password is incorrect', 'error')
        return redirect(url_for('profile'))
    
    # Update password
    new_hash = hashlib.sha256(new_password.encode()).hexdigest()
    cursor.execute("UPDATE users SET password_hash = %s WHERE id = %s", (new_hash, user_id))
    conn.commit()
    cursor.close()
    conn.close()
    
    flash('Password changed successfully!', 'success')
    return redirect(url_for('profile'))

@app.route('/update_profile', methods=['POST'])
def update_profile():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    new_password = request.form['new_password']
    if new_password:
        hash_pwd = hashlib.sha256(new_password.encode()).hexdigest()
        user_id = session['user_id']
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET password_hash = %s WHERE id = %s", (hash_pwd, user_id))
        conn.commit()
        cursor.close()
        conn.close()
        flash('Password updated successfully', 'success')
        
    return redirect(url_for('profile'))

@app.route('/delete_account', methods=['POST'])
def delete_account():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Fetch current profile picture filename before deleting user record
    cursor.execute("SELECT profile_pic FROM users WHERE id = %s", (user_id,))
    user = cursor.fetchone()
    profile_pic = user['profile_pic'] if user else None
    
    try:
        cursor.execute("DELETE FROM transactions WHERE user_id = %s", (user_id,))
        cursor.execute("DELETE FROM budgets WHERE user_id = %s", (user_id,))
        cursor.execute("DELETE FROM goals WHERE user_id = %s", (user_id,))
        cursor.execute("DELETE FROM notifications WHERE user_id = %s", (user_id,))
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        conn.commit()
        
        # Remove profile picture file if it's a custom upload (not the default placeholder)
        if profile_pic and profile_pic != 'default.png':
            pic_path = os.path.join(app.config['UPLOAD_FOLDER'], profile_pic)
            try:
                if os.path.isfile(pic_path):
                    os.remove(pic_path)
            except Exception as e:
                # Log the error silently; the deletion of the account is more important
                print(f"Failed to delete profile picture {pic_path}: {e}")
        
        session.clear()
        flash('Your account has been successfully deleted.', 'success')
        return redirect(url_for('register'))
    except mysql.connector.Error as err:
        conn.rollback()
        flash(f'Error deleting account: {err}', 'error')
        return redirect(url_for('profile'))
    finally:
        cursor.close()
        conn.close()

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check for goal deadlines and create notifications if needed
    try:
        check_goal_deadlines(session['user_id'])
    except Exception as e:
        print(f"Error checking goal deadlines: {e}")
    
    # Check balance status for notifications
    try:
        check_balance_status(session['user_id'])
    except Exception as e:
        print(f"Error checking balance status: {e}")
    
    return render_template('dashboard.html', username=session['username'], now=date.today())

# --- Transactions ---

@app.route('/transactions')
def transactions():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    # Determine selected month (YYYY-MM), default to current month
    selected_month = request.args.get('month')
    if not selected_month:
        selected_month = datetime.now().strftime('%Y-%m')
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Fetch non-deleted transactions for the selected month
    cursor.execute(
        "SELECT * FROM transactions WHERE user_id = %s AND (is_deleted IS NULL OR is_deleted = FALSE) AND DATE_FORMAT(date, '%Y-%m') = %s ORDER BY date DESC",
        (user_id, selected_month)
    )
    transactions = cursor.fetchall()
    
    # Calculate total spent for the selected month (all transactions, including soft-deleted for budget accuracy)
    cursor.execute(
        "SELECT SUM(amount) as total FROM transactions WHERE user_id = %s AND type = 'expense' AND DATE_FORMAT(date, '%Y-%m') = %s",
        (user_id, selected_month)
    )
    spent_res = cursor.fetchone()
    total_spent = spent_res['total'] or 0
    
    # Get total budget for the selected month
    cursor.execute(
        "SELECT SUM(limit_amount) as total FROM budgets WHERE user_id = %s AND month = %s",
        (user_id, selected_month)
    )
    budget_res = cursor.fetchone()
    total_budget = budget_res['total'] or 0
    
    remaining_budget = total_budget - total_spent
    
    # Fetch budgets for the selected month
    cursor.execute(
        "SELECT * FROM budgets WHERE user_id = %s AND month = %s",
        (user_id, selected_month)
    )
    budgets = cursor.fetchall()
    
    for b in budgets:
        cursor.execute(
            """
                SELECT COALESCE(SUM(amount), 0) as total FROM transactions 
                WHERE user_id = %s AND LOWER(category) = LOWER(%s) AND type = 'expense' 
                AND DATE_FORMAT(date, '%Y-%m') = %s
            """,
            (user_id, b['category'], selected_month)
        )
        res = cursor.fetchone()
        b['limit_amount'] = float(b['limit_amount'])  # Convert Decimal to float
        b['spent'] = float(res['total']) if res else 0.0
        b['remaining'] = b['limit_amount'] - b['spent']
        b['percent'] = (b['spent'] / b['limit_amount'] * 100) if b['limit_amount'] > 0 else 0
        b['remaining'] = float(b['limit_amount']) - float(b['spent'])
        b['percent'] = (float(b['spent']) / float(b['limit_amount']) * 100) if b['limit_amount'] > 0 else 0
    
    cursor.close()
    conn.close()
    
    return render_template(
        'transactions.html',
        transactions=transactions,
        budgets=budgets,
        today=date.today(),
        total_budget=total_budget,
        total_spent=total_spent,
        remaining_budget=remaining_budget,
        selected_month=selected_month
    )

@app.route('/add_transaction', methods=['POST'])
def add_transaction():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    tx_type = request.form['type']
    category = request.form['category'].strip()  # Strip whitespace
    amount = float(request.form['amount'])
    date_val = request.form['date']
    description = request.form.get('description', '')
    payment_method = request.form.get('payment_method', 'Cash')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    warning_msg = None

    try:
        # Insert the transaction
        cursor.execute(
            "INSERT INTO transactions (user_id, type, category, amount, description, date, payment_method, is_deleted) VALUES (%s, %s, %s, %s, %s, %s, %s, FALSE)",
            (user_id, tx_type, category, amount, description, date_val, payment_method)
        )
        conn.commit()
        
        # Check budget thresholds after adding expense
        if tx_type == 'expense':
            dt = datetime.strptime(date_val, '%Y-%m-%d')
            month_str = dt.strftime('%Y-%m')
            
            # Use a new cursor to ensure we see committed data
            cursor2 = conn.cursor(dictionary=True)
            
            # Check if budget exists for this category (case-insensitive)
            cursor2.execute(
                "SELECT id, limit_amount, category FROM budgets WHERE user_id = %s AND LOWER(category) = LOWER(%s) AND month = %s",
                (user_id, category, month_str)
            )
            budget = cursor2.fetchone()
            
            print(f"Looking for budget: user={user_id}, category='{category}', month={month_str}")
            print(f"Budget found: {budget}")
            
            if budget:
                limit = float(budget['limit_amount'])
                budget_category = budget['category']
                
                # Get total spent using the exact category from budget (include deleted for accurate tracking)
                cursor2.execute("""
                    SELECT COALESCE(SUM(amount), 0) as total FROM transactions 
                    WHERE user_id = %s AND LOWER(category) = LOWER(%s) AND type = 'expense' 
                    AND DATE_FORMAT(date, '%Y-%m') = %s
                """, (user_id, budget_category, month_str))
                
                result = cursor2.fetchone()
                current_spent = float(result['total']) if result else 0.0
                percent = (current_spent / limit * 100) if limit > 0 else 0
                
                print(f"Budget check: {category} - Spent: {current_spent}, Limit: {limit}, Percent: {percent:.1f}%")
                
                if percent >= 100:
                    warning_msg = f"Budget exceeded! You've spent Rs.{current_spent:.0f} of Rs.{limit:.0f} on {category}!"
                    cursor2.execute(
                        "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)", 
                        (user_id, warning_msg, 'danger')
                    )
                    conn.commit()
                    print(f"Notification created: {warning_msg}")
                elif percent >= 80:
                    warning_msg = f"Budget warning! You've used {percent:.0f}% of your {category} budget"
                    cursor2.execute(
                        "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)", 
                        (user_id, warning_msg, 'warning')
                    )
                    conn.commit()
                    print(f"Notification created: {warning_msg}")
            else:
                print(f"No budget found for category '{category}' in month {month_str}")
            
            cursor2.close()
            
            # Check for unusual spending pattern
            unusual_msg = check_unusual_spending(user_id, category, amount)
            if unusual_msg and not warning_msg:
                warning_msg = unusual_msg
        
        # Check balance after any transaction
        balance_warning = check_balance_status(user_id)
        if balance_warning and not warning_msg:
            warning_msg = balance_warning
        
        if warning_msg:
            flash(warning_msg, 'warning_sound')
        else:
            flash('Transaction added successfully!', 'success')
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        flash(f'Error adding transaction: {err}', 'error')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('transactions'))

@app.context_processor
def inject_notifications():
    if 'user_id' in session:
        user_id = session['user_id']
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM notifications WHERE user_id = %s ORDER BY date DESC LIMIT 10", (user_id,))
            notifs = cursor.fetchall()
            cursor.execute("SELECT COUNT(*) as count FROM notifications WHERE user_id = %s AND is_read = FALSE", (user_id,))
            res = cursor.fetchone()
            unread_count = res['count'] if res else 0
            return dict(notifications=notifs, unread_count=unread_count)
        except:
            return dict(notifications=[], unread_count=0)
        finally:
            conn.close()
    return dict(notifications=[], unread_count=0)

@app.route('/mark_read', methods=['POST'])
def mark_read():
    if 'user_id' in session:
        user_id = session['user_id']
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE notifications SET is_read = TRUE WHERE user_id = %s", (user_id,))
        conn.commit()
        cursor.close()
        conn.close()
    return '', 204

@app.route('/delete_notification/<int:id>', methods=['POST'])
def delete_notification(id):
    if 'user_id' in session:
        user_id = session['user_id']
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM notifications WHERE id = %s AND user_id = %s", (id, user_id))
        conn.commit()
        cursor.close()
        conn.close()
    return '', 204

@app.route('/clear_notifications', methods=['POST'])
def clear_notifications():
    if 'user_id' in session:
        user_id = session['user_id']
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM notifications WHERE user_id = %s", (user_id,))
        conn.commit()
        cursor.close()
        conn.close()
    return '', 204

@app.route('/api/set_theme', methods=['POST'])
def set_theme():
    """Set user's theme preference"""
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    theme = data.get('theme', 'dark')
    session['theme'] = theme
    return jsonify({'status': 'ok', 'theme': theme})

@app.route('/api/ping_session', methods=['POST'])
def ping_session():
    """Keep session alive when user is active"""
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    # Update last activity timestamp in session
    session['last_activity'] = datetime.now().isoformat()
    return jsonify({'status': 'ok'})

@app.route('/api/notifications')
def get_notifications():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM notifications WHERE user_id = %s ORDER BY date DESC LIMIT 20", (user_id,))
    notifs = cursor.fetchall()
    cursor.execute("SELECT COUNT(*) as count FROM notifications WHERE user_id = %s AND is_read = FALSE", (user_id,))
    res = cursor.fetchone()
    unread_count = res['count'] if res else 0
    cursor.close()
    conn.close()
    
    # Convert datetime objects to strings for JSON
    for n in notifs:
        if n.get('date'):
            n['date'] = n['date'].strftime('%Y-%m-%d %H:%M')
    
    return jsonify({'notifications': notifs, 'unread_count': unread_count})

# --- Notification Helper Functions ---

def create_notification(user_id, message, notif_type='info'):
    """Create a notification for a user"""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
            (user_id, message, notif_type)
        )
        conn.commit()
    except Exception as e:
        print(f"Error creating notification: {e}")
    finally:
        cursor.close()
        conn.close()

def get_current_balance(user_id):
    """Calculate user's current balance: initial_balance + income - expenses"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # Get initial balance
        cursor.execute("SELECT initial_balance FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        initial_balance = float(user['initial_balance']) if user and user['initial_balance'] else 0
        
        # Get total income
        cursor.execute("""
            SELECT COALESCE(SUM(amount), 0) as total FROM transactions 
            WHERE user_id = %s AND type = 'income' AND (is_deleted IS NULL OR is_deleted = FALSE)
        """, (user_id,))
        income_result = cursor.fetchone()
        total_income = float(income_result['total']) if income_result else 0
        
        # Get total expenses
        cursor.execute("""
            SELECT COALESCE(SUM(amount), 0) as total FROM transactions 
            WHERE user_id = %s AND type = 'expense' AND (is_deleted IS NULL OR is_deleted = FALSE)
        """, (user_id,))
        expense_result = cursor.fetchone()
        total_expenses = float(expense_result['total']) if expense_result else 0
        
        current_balance = initial_balance + total_income - total_expenses
        return current_balance
        
    except Exception as e:
        print(f"Error calculating balance: {e}")
        return 0
    finally:
        cursor.close()
        conn.close()

def check_balance_status(user_id):
    """Check if balance is low or negative and create notifications"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        current_balance = get_current_balance(user_id)
        
        # Check if notification was already sent today for this condition
        if current_balance < 0:
            # Negative balance - critical warning
            cursor.execute("""
                SELECT id FROM notifications 
                WHERE user_id = %s AND message LIKE '%negative%' AND DATE(date) = CURDATE()
            """, (user_id,))
            
            if not cursor.fetchone():
                msg = f"🚨 Alert! Your balance is negative: ₹{current_balance:,.2f}. Please add income or review expenses."
                cursor.execute(
                    "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
                    (user_id, msg, 'danger')
                )
                conn.commit()
                return msg
                
        elif current_balance < 1000:
            # Low balance warning (below ₹1000)
            cursor.execute("""
                SELECT id FROM notifications 
                WHERE user_id = %s AND message LIKE '%Low balance%' AND DATE(date) = CURDATE()
            """, (user_id,))
            
            if not cursor.fetchone():
                msg = f"⚠️ Low balance alert: Only ₹{current_balance:,.2f} remaining. Consider reducing expenses."
                cursor.execute(
                    "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
                    (user_id, msg, 'warning')
                )
                conn.commit()
                return msg
        
        return None
        
    except Exception as e:
        print(f"Error checking balance status: {e}")
        return None
    finally:
        cursor.close()
        conn.close()

def check_budget_thresholds(user_id, category, month_str):
    """Check budget thresholds and create notifications if needed"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # Get budget limit
        cursor.execute(
            "SELECT limit_amount FROM budgets WHERE user_id = %s AND category = %s AND month = %s",
            (user_id, category, month_str)
        )
        budget = cursor.fetchone()
        
        if not budget:
            return
        
        limit = float(budget['limit_amount'])
        
        # Get current spending
        cursor.execute("""
            SELECT SUM(amount) as total FROM transactions 
            WHERE user_id = %s AND category = %s AND type = 'expense' 
            AND DATE_FORMAT(date, '%Y-%m') = %s
        """, (user_id, category, month_str))
        
        result = cursor.fetchone()
        spent = float(result['total']) if result and result['total'] else 0
        percent = (spent / limit * 100) if limit > 0 else 0
        
        # Check thresholds and avoid duplicate notifications
        if percent >= 100:
            msg = f"🚨 Budget exceeded! You've spent ₹{spent:.0f} of ₹{limit:.0f} ({percent:.0f}%) on {category}"
            notif_type = 'danger'
        elif percent >= 80:
            msg = f"⚠️ Budget warning! You've used {percent:.0f}% of your {category} budget (₹{spent:.0f}/₹{limit:.0f})"
            notif_type = 'warning'
        elif percent >= 50:
            msg = f"📊 Heads up: You've used {percent:.0f}% of your {category} budget"
            notif_type = 'info'
        else:
            return
        
        # Check if similar notification exists today
        cursor.execute("""
            SELECT id FROM notifications 
            WHERE user_id = %s AND message LIKE %s AND DATE(date) = CURDATE()
        """, (user_id, f"%{category}%{notif_type}%"))
        
        if not cursor.fetchone():
            cursor.execute(
                "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
                (user_id, msg, notif_type)
            )
            conn.commit()
            
    except Exception as e:
        print(f"Error checking budget thresholds: {e}")
    finally:
        cursor.close()
        conn.close()

def check_unusual_spending(user_id, category, amount):
    """Check if a transaction is unusually high compared to average spending in that category"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # Get average expense amount for this category (last 3 months)
        cursor.execute("""
            SELECT AVG(amount) as avg_amount, COUNT(*) as count, MAX(amount) as max_amount
            FROM transactions 
            WHERE user_id = %s AND LOWER(category) = LOWER(%s) AND type = 'expense'
            AND date >= DATE_SUB(CURDATE(), INTERVAL 3 MONTH)
        """, (user_id, category))
        
        result = cursor.fetchone()
        
        if result and result['count'] and result['count'] >= 3:
            avg_amount = float(result['avg_amount']) if result['avg_amount'] else 0
            max_amount = float(result['max_amount']) if result['max_amount'] else 0
            
            # Alert if transaction is more than 2x the average or exceeds previous max by 50%
            if avg_amount > 0:
                if amount >= avg_amount * 2.5:
                    msg = f"⚠️ Unusual spending detected! ₹{amount:.0f} on {category} is {(amount/avg_amount):.1f}x your average (₹{avg_amount:.0f})"
                    cursor.execute(
                        "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
                        (user_id, msg, 'warning')
                    )
                    conn.commit()
                    return msg
                elif max_amount > 0 and amount >= max_amount * 1.5:
                    msg = f"📊 Large expense alert: ₹{amount:.0f} on {category} exceeds your usual spending pattern"
                    cursor.execute(
                        "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
                        (user_id, msg, 'info')
                    )
                    conn.commit()
                    return msg
        
        # Also check for overall daily spending being unusual
        cursor.execute("""
            SELECT AVG(daily_total) as avg_daily FROM (
                SELECT DATE(date) as tx_date, SUM(amount) as daily_total
                FROM transactions 
                WHERE user_id = %s AND type = 'expense'
                AND date >= DATE_SUB(CURDATE(), INTERVAL 1 MONTH)
                GROUP BY DATE(date)
            ) daily_totals
        """, (user_id,))
        
        daily_result = cursor.fetchone()
        
        if daily_result and daily_result['avg_daily']:
            avg_daily = float(daily_result['avg_daily'])
            
            # Get today's total spending
            cursor.execute("""
                SELECT SUM(amount) as today_total FROM transactions
                WHERE user_id = %s AND type = 'expense' AND DATE(date) = CURDATE()
            """, (user_id,))
            
            today_result = cursor.fetchone()
            today_total = float(today_result['today_total']) if today_result and today_result['today_total'] else 0
            
            # Check if similar alert was already sent today
            cursor.execute("""
                SELECT id FROM notifications 
                WHERE user_id = %s AND message LIKE %s AND DATE(date) = CURDATE()
            """, (user_id, "%daily spending%"))
            
            if not cursor.fetchone() and today_total >= avg_daily * 2:
                msg = f"💰 High daily spending: You've spent ₹{today_total:.0f} today, which is {(today_total/avg_daily):.1f}x your daily average"
                cursor.execute(
                    "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
                    (user_id, msg, 'warning')
                )
                conn.commit()
                return msg
                
    except Exception as e:
        print(f"Error checking unusual spending: {e}")
    finally:
        cursor.close()
        conn.close()
    
    return None

def check_goal_deadlines(user_id):
    """Check for upcoming goal deadlines and notify user"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # Get goals with deadlines in next 7 days
        cursor.execute("""
            SELECT * FROM goals 
            WHERE user_id = %s AND deadline IS NOT NULL 
            AND deadline BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 7 DAY)
            AND current_amount < target_amount
        """, (user_id,))
        
        goals = cursor.fetchall()
        
        for goal in goals:
            days_left = (goal['deadline'] - date.today()).days
            remaining = float(goal['target_amount']) - float(goal['current_amount'])
            
            if days_left <= 1:
                msg = f"🎯 Goal '{goal['name']}' deadline is tomorrow! ₹{remaining:.0f} still needed."
                notif_type = 'danger'
            elif days_left <= 3:
                msg = f"⏰ Goal '{goal['name']}' deadline in {days_left} days. ₹{remaining:.0f} remaining."
                notif_type = 'warning'
            else:
                msg = f"📅 Goal '{goal['name']}' deadline approaching in {days_left} days."
                notif_type = 'info'
            
            # Check for existing notification
            cursor.execute("""
                SELECT id FROM notifications 
                WHERE user_id = %s AND message LIKE %s AND DATE(date) = CURDATE()
            """, (user_id, f"%{goal['name']}%deadline%"))
            
            if not cursor.fetchone():
                cursor.execute(
                    "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
                    (user_id, msg, notif_type)
                )
                conn.commit()
                
    except Exception as e:
        print(f"Error checking goal deadlines: {e}")
    finally:
        cursor.close()
        conn.close()

@app.route('/delete_transaction/<int:id>', methods=['POST'])
def delete_transaction(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE transactions SET is_deleted = TRUE WHERE id = %s AND user_id = %s", (id, user_id))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Transaction removed', 'success')
    return redirect(url_for('transactions'))

# --- Budget ---

@app.route('/budget')
def budget():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    current_month = datetime.now().strftime('%Y-%m')
    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Delete budgets from previous months to reset them
    cursor.execute("DELETE FROM budgets WHERE user_id = %s AND month < %s", (user_id, current_month))
    conn.commit()
    
    # Fetch budgets for the current month only
    cursor.execute("SELECT * FROM budgets WHERE user_id = %s AND month = %s", (user_id, current_month))
    budgets = cursor.fetchall()
    
    for b in budgets:
        cursor.execute("""
            SELECT COALESCE(SUM(amount), 0) as total FROM transactions 
            WHERE user_id = %s AND LOWER(category) = LOWER(%s) AND type = 'expense' 
            AND DATE_FORMAT(date, '%Y-%m') = %s
        """, (user_id, b['category'], current_month))
        result = cursor.fetchone()
        b['limit_amount'] = float(b['limit_amount'])  # Convert Decimal to float
        b['spent'] = float(result['total']) if result else 0.0
        b['remaining'] = b['limit_amount'] - b['spent']
        b['percent'] = (b['spent'] / b['limit_amount'] * 100) if b['limit_amount'] > 0 else 0
    
    cursor.close()
    conn.close()
    
    return render_template('budget.html', budgets=budgets, current_month=current_month)

@app.route('/add_budget', methods=['POST'])
def add_budget():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    category = request.form['category']
    limit = request.form['limit_amount']
    month = request.form['month']
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM budgets WHERE user_id = %s AND category = %s AND month = %s", (user_id, category, month))
    if cursor.fetchone():
        flash('Budget for this category already exists!', 'error')
    else:
        cursor.execute("INSERT INTO budgets (user_id, category, limit_amount, month) VALUES (%s, %s, %s, %s)",
                       (user_id, category, limit, month))
        conn.commit()
        
        # Create notification for new budget
        msg = f"💰 Budget set: ₹{float(limit):.0f} for {category} in {month}"
        try:
            cursor.execute("INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, 'info')",
                          (user_id, msg))
            conn.commit()
        except:
            pass
        
        flash('Budget set successfully', 'success')
    cursor.close()
    conn.close()
    return redirect(url_for('budget'))

@app.route('/update_budget', methods=['POST'])
def update_budget():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    budget_id = request.form['budget_id']
    limit = request.form['limit_amount']
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE budgets SET limit_amount = %s WHERE id = %s AND user_id = %s",
                   (limit, budget_id, session['user_id']))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Budget updated successfully', 'success')
    return redirect(url_for('budget'))

@app.route('/delete_budget/<int:id>', methods=['POST'])
def delete_budget(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM budgets WHERE id = %s AND user_id = %s", (id, session['user_id']))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('budget'))

# --- Goals ---

@app.route('/goals')
def goals():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM goals WHERE user_id = %s", (user_id,))
    goals = cursor.fetchall()
    total_target = sum(g['target_amount'] for g in goals)
    total_saved = sum(g['current_amount'] for g in goals)
    cursor.close()
    conn.close()
    return render_template('goals.html', goals=goals, total_target=total_target, total_saved=total_saved)

@app.route('/add_goal', methods=['POST'])
def add_goal():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']
    name = request.form['name']
    target = request.form['target_amount']
    current = request.form.get('current_amount', 0)
    deadline = request.form['deadline'] or None
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO goals (user_id, name, target_amount, current_amount, deadline) VALUES (%s, %s, %s, %s, %s)",
                   (user_id, name, target, current, deadline))
    conn.commit()
    
    # Create notification for new goal
    deadline_str = f" by {deadline}" if deadline else ""
    msg = f"🎯 New goal created: '{name}' - Target ₹{float(target):.0f}{deadline_str}"
    try:
        cursor.execute("INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, 'info')",
                      (user_id, msg))
        conn.commit()
    except:
        pass

    # If initial amount > 0, record as transaction
    if float(current) > 0:
        cursor.execute("""
            INSERT INTO transactions (user_id, type, category, amount, date, description, payment_method)
            VALUES (%s, 'expense', 'Financial Goal', %s, CURDATE(), %s, 'Savings')
        """, (user_id, current, f"Initial deposit for goal: {name}"))
        conn.commit()
    
    cursor.close()
    conn.close()
    flash('Goal added successfully', 'success')
    return redirect(url_for('goals'))

@app.route('/update_goal', methods=['POST'])
def update_goal():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    goal_id = request.form['goal_id']
    amount = float(request.form['amount'])
    user_id = session['user_id']
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get current goal state
    cursor.execute("SELECT * FROM goals WHERE id = %s AND user_id = %s", (goal_id, user_id))
    goal = cursor.fetchone()
    
    if goal:
        new_amount = float(goal['current_amount']) + amount
        target = float(goal['target_amount'])
        
        cursor.execute("UPDATE goals SET current_amount = %s WHERE id = %s", (new_amount, goal_id))
        conn.commit()
        
        # Record transaction
        tx_type = 'expense' if amount > 0 else 'income'
        tx_desc = f"Added to goal: {goal['name']}" if amount > 0 else f"Withdrawn from goal: {goal['name']}"
        abs_amount = abs(amount)
        
        cursor.execute("""
            INSERT INTO transactions (user_id, type, category, amount, date, description, payment_method)
            VALUES (%s, %s, 'Financial Goal', %s, CURDATE(), %s, 'Savings')
        """, (user_id, tx_type, abs_amount, tx_desc))
        conn.commit()
        
        # Check if goal is now complete
        if new_amount >= target and float(goal['current_amount']) < target:
            msg = f"🎉 Congratulations! You've reached your goal '{goal['name']}'! Target: ₹{target:.0f}"
            cursor.execute("INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, 'success')",
                          (user_id, msg))
            conn.commit()
            flash(f'Goal completed! 🎉', 'success')
        else:
            flash('Goal updated', 'success')
    
    cursor.close()
    conn.close()
    return redirect(url_for('goals'))

@app.route('/delete_goal/<int:id>', methods=['POST'])
def delete_goal(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM goals WHERE id = %s AND user_id = %s", (id, session['user_id']))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('goals'))

# --- Bills ---

@app.route('/bills')
def bills():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    today = date.today()
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Check for bill reminders and create notifications
    check_bill_reminders(user_id)
    
    # Get all bills for the user
    cursor.execute("""
        SELECT * FROM bills WHERE user_id = %s ORDER BY is_paid ASC, due_date ASC
    """, (user_id,))
    all_bills = cursor.fetchall()
    
    # Process bills to add status info
    pending_bills = []
    overdue_count = 0
    due_soon_count = 0
    total_pending = 0
    
    for bill in all_bills:
        days_until = (bill['due_date'] - today).days
        bill['days_until'] = days_until
        bill['is_overdue'] = not bill['is_paid'] and days_until < 0
        bill['is_due_soon'] = not bill['is_paid'] and 0 <= days_until <= 3
        bill['amount'] = float(bill['amount'])
        
        if not bill['is_paid']:
            pending_bills.append(bill)
            total_pending += bill['amount']
            if bill['is_overdue']:
                overdue_count += 1
            elif bill['is_due_soon']:
                due_soon_count += 1
    
    cursor.close()
    conn.close()
    
    return render_template('bills.html', 
                           bills=all_bills, 
                           pending_bills=pending_bills,
                           overdue_count=overdue_count,
                           due_soon_count=due_soon_count,
                           total_pending=total_pending,
                           today=today)

def check_bill_reminders(user_id):
    """Check bills and create notifications for due/overdue bills"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    today = date.today()
    
    # Get unpaid bills
    cursor.execute("""
        SELECT * FROM bills WHERE user_id = %s AND is_paid = FALSE
    """, (user_id,))
    bills = cursor.fetchall()
    
    for bill in bills:
        days_until = (bill['due_date'] - today).days
        
        # Check if notification already exists today for this bill
        cursor.execute("""
            SELECT id FROM notifications 
            WHERE user_id = %s AND message LIKE %s AND DATE(date) = %s
        """, (user_id, f"%{bill['name']}%", today))
        
        existing = cursor.fetchall()  # Consume all results
        if existing:
            continue  # Skip if already notified today
        
        if days_until < 0:
            # Overdue
            msg = f"🚨 OVERDUE: {bill['name']} (₹{float(bill['amount']):.0f}) was due {-days_until} days ago!"
            cursor.execute("INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, 'danger')",
                          (user_id, msg))
            conn.commit()
        elif days_until == 0:
            # Due today
            msg = f"⏰ DUE TODAY: {bill['name']} (₹{float(bill['amount']):.0f}) is due today!"
            cursor.execute("INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, 'warning')",
                          (user_id, msg))
            conn.commit()
        elif days_until <= 3:
            # Due soon
            msg = f"📅 REMINDER: {bill['name']} (₹{float(bill['amount']):.0f}) is due in {days_until} days"
            cursor.execute("INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, 'info')",
                          (user_id, msg))
            conn.commit()
    
    cursor.close()
    conn.close()

@app.route('/add_bill', methods=['POST'])
def add_bill():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    name = request.form['name']
    amount = float(request.form['amount'])
    due_date = request.form['due_date']
    category = request.form['category']
    is_recurring = 'is_recurring' in request.form
    recurrence = request.form.get('recurrence', 'monthly')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        INSERT INTO bills (user_id, name, amount, due_date, category, is_recurring, recurrence)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (user_id, name, amount, due_date, category, is_recurring, recurrence))
    conn.commit()
    
    # Create notification for new bill
    cursor.execute("""
        INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, 'info')
    """, (user_id, f"📝 Bill added: {name} (₹{amount:.0f}) due on {due_date}"))
    conn.commit()
    
    cursor.close()
    conn.close()
    
    flash('Bill added successfully!', 'success')
    return redirect(url_for('bills'))

@app.route('/mark_bill_paid/<int:id>', methods=['POST'])
def mark_bill_paid(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    today = date.today()
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get bill details
    cursor.execute("SELECT * FROM bills WHERE id = %s AND user_id = %s", (id, user_id))
    bill = cursor.fetchone()
    
    if bill:
        # Mark as paid
        cursor.execute("UPDATE bills SET is_paid = TRUE, paid_date = %s WHERE id = %s", (today, id))
        conn.commit()
        
        # If recurring, create next bill
        if bill['is_recurring']:
            from dateutil.relativedelta import relativedelta
            
            next_due = bill['due_date']
            if bill['recurrence'] == 'weekly':
                next_due += relativedelta(weeks=1)
            elif bill['recurrence'] == 'monthly':
                next_due += relativedelta(months=1)
            elif bill['recurrence'] == 'yearly':
                next_due += relativedelta(years=1)
            
            cursor.execute("""
                INSERT INTO bills (user_id, name, amount, due_date, category, is_recurring, recurrence)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (user_id, bill['name'], bill['amount'], next_due, bill['category'], True, bill['recurrence']))
            conn.commit()
            
            flash(f'Bill paid! Next {bill["recurrence"]} bill created for {next_due}', 'success')
        else:
            flash('Bill marked as paid!', 'success')
        
        # Create notification
        cursor.execute("""
            INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, 'success')
        """, (user_id, f"✅ Bill paid: {bill['name']} (₹{float(bill['amount']):.0f})"))
        conn.commit()
    
    cursor.close()
    conn.close()
    
    return redirect(url_for('bills'))

@app.route('/mark_bill_unpaid/<int:id>', methods=['POST'])
def mark_bill_unpaid(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE bills SET is_paid = FALSE, paid_date = NULL WHERE id = %s AND user_id = %s", 
                   (id, session['user_id']))
    conn.commit()
    cursor.close()
    conn.close()
    
    flash('Bill marked as unpaid', 'info')
    return redirect(url_for('bills'))

@app.route('/delete_bill/<int:id>', methods=['POST'])
def delete_bill(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM bills WHERE id = %s AND user_id = %s", (id, session['user_id']))
    conn.commit()
    cursor.close()
    conn.close()
    
    flash('Bill deleted', 'success')
    return redirect(url_for('bills'))

# --- Reports ---

@app.route('/report')
def report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('report.html')

@app.route('/download_report/<type>')
def download_report(type):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    conn = get_db_connection()

    # ---- Transactions ----
    query = "SELECT date, type, category, amount, description FROM transactions WHERE user_id = %s ORDER BY date DESC"
    df_transactions = pd.read_sql(query, conn, params=(user_id,))

    # ---- Budgets ----
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, category, limit_amount, month FROM budgets WHERE user_id = %s", (user_id,))
    budgets = cursor.fetchall()
    for b in budgets:
        cursor.execute(
            """
                SELECT SUM(amount) as total FROM transactions 
                WHERE user_id = %s AND category = %s AND type = 'expense' 
                AND DATE_FORMAT(date, '%Y-%m') = %s
            """,
            (user_id, b['category'], b['month'])
        )
        res = cursor.fetchone()
        spent = res['total'] or 0
        b['spent'] = spent
        b['remaining'] = float(b['limit_amount']) - float(spent)
        b['percent'] = (float(spent) / float(b['limit_amount']) * 100) if b['limit_amount'] > 0 else 0
    
    df_budgets = pd.DataFrame(budgets)
    if not df_budgets.empty:
        df_budgets = df_budgets[['category', 'limit_amount', 'spent', 'remaining', 'percent', 'month']]
        df_budgets.columns = ['Category', 'Limit', 'Spent', 'Remaining', 'Percent', 'Month']

    # ---- Bills ----
    cursor.execute("""
        SELECT name, amount, due_date, category, is_recurring, recurrence, is_paid, paid_date 
        FROM bills WHERE user_id = %s ORDER BY due_date
    """, (user_id,))
    bills = cursor.fetchall()
    df_bills = pd.DataFrame(bills)
    if not df_bills.empty:
        df_bills.columns = ['Bill Name', 'Amount', 'Due Date', 'Category', 'Recurring', 'Frequency', 'Paid', 'Paid Date']
    
    cursor.close()
    conn.close()

    if type == 'excel':
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_transactions.to_excel(writer, index=False, sheet_name='Transactions')
            if not df_budgets.empty:
                df_budgets.to_excel(writer, index=False, sheet_name='Budgets')
            if not df_bills.empty:
                df_bills.to_excel(writer, index=False, sheet_name='Bills')
        output.seek(0)
        return send_file(output, download_name='finance_report.xlsx', as_attachment=True)
    elif type == 'pdf':
        class PDF(FPDF):
            def header(self):
                # Accent Color Line
                self.set_fill_color(114, 105, 227) # Primary
                self.rect(0, 0, 210, 5, 'F')
                
                self.set_y(15)
                # Logo
                try:
                    logo_path = os.path.join(app.root_path, 'static', 'logo.png')
                    if os.path.exists(logo_path):
                        self.image(logo_path, 10, 10, 30)
                except: 
                    pass
                
                self.set_font('Arial', 'B', 20)
                self.set_text_color(30, 41, 59)
                self.cell(0, 10, 'Finance Report', 0, 1, 'C')
                
                self.set_font('Arial', 'I', 10)
                self.set_text_color(100, 116, 139)
                self.cell(0, 5, f'Generated on {date.today().strftime("%B %d, %Y")}', 0, 1, 'C')
                self.ln(15)

            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.set_text_color(148, 163, 184)
                self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

            def section_title(self, title):
                self.set_font('Arial', 'B', 14)
                self.set_text_color(114, 105, 227)
                self.cell(0, 10, title, 0, 1, 'L')
                self.line(10, self.get_y(), 200, self.get_y())
                self.ln(5)

            def table_header(self, headers, widths):
                self.set_font('Arial', 'B', 10)
                self.set_fill_color(241, 245, 249)
                self.set_text_color(51, 65, 85)
                for i, h in enumerate(headers):
                    self.cell(widths[i], 10, h, 0, 0, 'L', True)
                self.ln()

            def table_row(self, data, widths, fill=False):
                self.set_font('Arial', '', 9)
                self.set_fill_color(248, 250, 252)
                self.set_text_color(51, 65, 85)
                for i, d in enumerate(data):
                    self.cell(widths[i], 9, str(d), 0, 0, 'L', fill)
                self.ln()

        pdf = PDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        
        # Summary Section
        income = df_transactions[df_transactions['type']=='income']['amount'].sum()
        expense = df_transactions[df_transactions['type']=='expense']['amount'].sum()
        savings = income - expense
        
        pdf.set_font('Arial', 'B', 12)
        pdf.set_fill_color(241, 245, 249)
        pdf.cell(63, 20, f"Income: +Rs {income:,.2f}", 0, 0, 'C', True)
        pdf.cell(1, 20, "", 0, 0) # Spacer
        pdf.cell(63, 20, f"Expense: -Rs {expense:,.2f}", 0, 0, 'C', True)
        pdf.cell(1, 20, "", 0, 0) # Spacer
        pdf.cell(63, 20, f"Net: Rs {savings:,.2f}", 0, 1, 'C', True)
        pdf.ln(10)

        # Transactions
        pdf.section_title("Recent Transactions")
        cols = ['Date', 'Cat', 'Description', 'Type', 'Amount']
        widths = [25, 30, 75, 20, 40]
        pdf.table_header(cols, widths)
        
        fill = False
        for _, row in df_transactions.iterrows():
            amt_str = f"Rs {row['amount']:,.2f}"
            desc = (row['description'][:35] + '..') if len(row['description']) > 35 else row['description']
            data = [str(row['date']), row['category'], desc, row['type'].title(), amt_str]
            pdf.table_row(data, widths, fill)
            fill = not fill # Toggle row color

        # Budgets
        if not df_budgets.empty:
            pdf.add_page()
            pdf.section_title("Budget Status")
            cols = ['Category', 'Limit', 'Spent', 'Remaining', '%']
            widths = [50, 35, 35, 35, 35]
            pdf.table_header(cols, widths)
            fill = False
            for _, row in df_budgets.iterrows():
                data = [
                    row['Category'], 
                    f"Rs {row['Limit']:,.0f}", 
                    f"Rs {row['Spent']:,.0f}",
                    f"Rs {row['Remaining']:,.0f}",
                    f"{row['Percent']:.1f}%"
                ]
                pdf.table_row(data, widths, fill)
                fill = not fill

        # Bills
        if not df_bills.empty:
            if pdf.get_y() > 200: pdf.add_page()
            else: pdf.ln(10)
            
            pdf.section_title("Upcoming Bills")
            cols = ['Bill', 'Due Date', 'Amount', 'Status']
            widths = [60, 40, 40, 50]
            pdf.table_header(cols, widths)
            fill = False
            for _, row in df_bills.iterrows():
                status = "Paid" if row['Paid'] else "Pending"
                data = [
                    row['Bill Name'],
                    str(row['Due Date']),
                    f"Rs {float(row['Amount']):,.2f}",
                    status
                ]
                pdf.table_row(data, widths, fill)
                fill = not fill

        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            pdf.output(tmp.name)
            tmp_path = tmp.name
        return send_file(tmp_path, download_name='finance_report.pdf', as_attachment=True)
    return redirect(url_for('report'))
def download_report(type):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    conn = get_db_connection()
    
    query = "SELECT date, type, category, amount, description FROM transactions WHERE user_id = %s ORDER BY date DESC"
    df = pd.read_sql(query, conn, params=(user_id,))
    conn.close()
    
    if type == 'excel':
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Transactions')
        output.seek(0)
        return send_file(output, download_name='finance_report.xlsx', as_attachment=True)
    elif type == 'pdf':
        class PDF(FPDF):
            def header(self):
                try:
                    logo_path = os.path.join(app.root_path, 'static', 'logo.png')
                    if os.path.exists(logo_path):
                        img_width = 150
                        x = (self.w - img_width) / 2
                        y = (self.h - img_width) / 2
                        self.image(logo_path, x=x, y=y, w=img_width)
                except Exception as e:
                    print(f"Watermark error: {e}")
                self.set_font('Arial', 'B', 15)
                self.cell(0, 10, 'Finance Report', 0, 1, 'C')
                self.ln(10)
            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')
        pdf = PDF()
        pdf.add_page()
        pdf.set_font("Arial", size=10)
        cols = ['Date', 'Type', 'Category', 'Amount', 'Description']
        col_widths = [30, 20, 40, 30, 70]
        for i, col in enumerate(cols):
            pdf.cell(col_widths[i], 10, col, 1, 0, 'C')
        pdf.ln()
        for idx, row in df.iterrows():
            pdf.cell(col_widths[0], 10, str(row['date']), 1)
            pdf.cell(col_widths[1], 10, row['type'], 1)
            pdf.cell(col_widths[2], 10, str(row['category']), 1)
            pdf.cell(col_widths[3], 10, str(row['amount']), 1)
            pdf.cell(col_widths[4], 10, str(row['description'])[:30], 1)
            pdf.ln()
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            pdf.output(tmp.name)
            tmp_path = tmp.name
        return send_file(tmp_path, download_name='finance_report.pdf', as_attachment=True)
    return redirect(url_for('report'))

# --- API ---

@app.route('/api/dashboard_data')
def dashboard_data():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get initial balance
    cursor.execute("SELECT initial_balance FROM users WHERE id = %s", (user_id,))
    user_data = cursor.fetchone()
    initial_balance = float(user_data['initial_balance']) if user_data and user_data['initial_balance'] else 0
    
    cursor.execute("SELECT SUM(amount) as total FROM transactions WHERE user_id = %s AND type = 'income'", (user_id,))
    income = cursor.fetchone()['total'] or 0
    
    cursor.execute("SELECT SUM(amount) as total FROM transactions WHERE user_id = %s AND type = 'expense'", (user_id,))
    expense = cursor.fetchone()['total'] or 0
    
    # Balance now includes initial balance
    balance = initial_balance + float(income) - float(expense)
    
    cursor.execute("SELECT * FROM transactions WHERE user_id = %s AND (is_deleted IS NULL OR is_deleted = FALSE) ORDER BY date DESC LIMIT 5", (user_id,))
    recent_transactions = cursor.fetchall()
    
    cursor.execute("SELECT category, SUM(amount) as total FROM transactions WHERE user_id = %s AND type = 'expense' GROUP BY category", (user_id,))
    category_data = cursor.fetchall()
    
    # Monthly aggregates for last 12 months (including soft-deleted)
    cursor.execute("""
        SELECT DATE_FORMAT(date, '%Y-%m') as month,
               SUM(CASE WHEN type='income' THEN amount ELSE 0 END) as income,
               SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) as expense
        FROM transactions
        WHERE user_id = %s
        GROUP BY month
        ORDER BY month DESC
        LIMIT 12
    """, (user_id,))
    monthly = cursor.fetchall()
    
    today = date.today()
    
    # Optimized Bills Query
    cursor.execute("SELECT * FROM bills WHERE user_id = %s AND is_paid = FALSE ORDER BY due_date ASC", (user_id,))
    bills = cursor.fetchall()
    
    pending_bills = len(bills)
    overdue_bills = 0
    due_soon = 0
    pending_amount = 0
    upcoming_bills = []
    
    for bill in bills:
        days_until = (bill['due_date'] - today).days
        bill['is_overdue'] = days_until < 0
        bill['is_due_soon'] = 0 <= days_until <= 3
        
        if bill['is_overdue']:
            overdue_bills += 1
        elif bill['is_due_soon']:
            due_soon += 1
            
        pending_amount += float(bill['amount'])
        
        # Add to upcoming list (limit 5)
        if len(upcoming_bills) < 5:
            # Convert decimal to float for JSON serialization
            bill['amount'] = float(bill['amount'])
            if isinstance(bill['due_date'], (date, datetime)):
                bill['due_date'] = bill['due_date'].strftime('%Y-%m-%d')
            upcoming_bills.append(bill)

    # Get goals
    cursor.execute("SELECT * FROM goals WHERE user_id = %s", (user_id,))
    goals = cursor.fetchall()
    processed_goals = []
    for goal in goals:
        # Calculate percentage
        target = float(goal['target_amount'])
        current = float(goal['current_amount'])
        percent = (current / target * 100) if target > 0 else 0
        
        processed_goals.append({
            'name': goal['name'],
            'target_amount': target,
            'current_amount': current,
            'percentage': round(percent, 1),
            'deadline': goal['deadline'].strftime('%Y-%m-%d') if goal['deadline'] else None
        })
    
    cursor.close()
    conn.close()
    
    return jsonify({
        'income': float(income),
        'expense': float(expense),
        'balance': float(balance),
        'transactions': recent_transactions,
        'categories': category_data,
        'monthly': monthly,
        'bills': {
            'pending_count': pending_bills,
            'overdue_count': overdue_bills,
            'due_soon_count': due_soon,
            'pending_amount': float(pending_amount),
            'upcoming': upcoming_bills
        },
        'goals': processed_goals
    })

if __name__ == '__main__':
    app.run(debug=True, port=5000)
